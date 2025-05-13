import math
from flask import Blueprint, render_template, jsonify, request, send_file
from pymongo import MongoClient
import os
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
from bson import json_util
import json
from datetime import datetime
from io import BytesIO
import openpyxl  # Alternative to xlsxwriter
import numpy as np
from plotly.subplots import make_subplots  # Import for make_subplots


production_overview = Blueprint('production_overview', __name__)

# MongoDB Atlas connection string
MONGO_URI = os.environ.get('MONGO_URI', 'mongodb+srv://nikhilvjagtap2003:ra9be97kyhd8wsR9@cluster0.vrqby.mongodb.net/quality_dashboard?retryWrites=true&w=majority')

try:
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
    db = client.get_database("quality_dashboard")
except Exception as e:
    print(f"Failed to connect to MongoDB: {e}")
    db = None

@production_overview.route('/production_overview')
def show_production_overview():
    if db is None:
        return jsonify({'error': 'Database connection error'}), 500
        
    try:
        return render_template('production_overview.html')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/summary_metrics')
def get_summary_metrics():
    if db is None:
        return jsonify({'error': 'Database connection error'}), 500
        
    try:
        # Fetch all coil data from MongoDB
        coil_data = list(db.coil_production.find({}, {'_id': 0}))
        
        # Convert to DataFrame
        df = pd.DataFrame(coil_data)
        
        # Calculate metrics
        total_coils = len(df)
        avg_thickness = df['finalThk'].mean() if 'finalThk' in df.columns else None
        avg_width = df['finalWidth'].mean() if 'finalWidth' in df.columns else None
        
        # Calculate approval rate
        if 'labResult' in df.columns:
            approved_count = df[df['labResult'] == 'Pass'].shape[0]
            approval_rate = (approved_count / total_coils) * 100 if total_coils > 0 else 0
        else:
            approval_rate = None
        
        return jsonify({
            'totalCoils': total_coils,
            'avgThickness': avg_thickness,
            'avgWidth': avg_width,
            'approvalRate': approval_rate
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/daily_production_chart')
def get_daily_production_chart():
  if db is None:
      return jsonify({'error': 'Database connection error'}), 500
      
  try:
      # Fetch coil data
      coil_data = list(db.coil_production.find({}, {'_id': 0, 'clearanceDate': 1}))
      
      # Process data for chart
      df = pd.DataFrame(coil_data)
      
      # If no data, return empty result
      if df.empty:
          return jsonify({'data': [], 'labels': []})
      
      # Convert clearanceDate to datetime
      df['clearanceDate'] = pd.to_datetime(df['clearanceDate'])
      
      # Extract date from clearanceDate (without time)
      df['production_date'] = df['clearanceDate'].dt.date
      
      # Group by date and count coils
      daily_counts = df.groupby('production_date').size().reset_index(name='count')
      daily_counts = daily_counts.sort_values('production_date')
      
      # Convert dates to strings for JSON serialization
      dates = [d.strftime('%Y-%m-%d') for d in daily_counts['production_date']]
      counts = daily_counts['count'].tolist()
      
      return jsonify({
          'labels': dates,
          'data': counts
      })
  except Exception as e:
      return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/coil_analysis_chart')
def get_coil_analysis_chart():
  if db is None:
      return jsonify({'error': 'Database connection error'}), 500
      
  try:
      # Get filter parameters
      from_date = request.args.get('from_date')
      to_date = request.args.get('to_date')
      product = request.args.get('product')
      
      # Build query
      query = {}
      if from_date and to_date:
          query['clearanceDate'] = {
              '$gte': from_date,
              '$lte': to_date
          }
      if product:
          query['product'] = product
      
      # Fetch filtered coil data
      coil_data = list(db.coil_production.find(query, {
          '_id': 0,
          'clearanceDate': 1,
          'planThickness': 1,
          'finalThk': 1,
          'planWidth': 1,
          'finalWidth': 1,
          'finalWeight': 1,
          'product': 1
      }))
      
      # Process data for chart
      df = pd.DataFrame(coil_data)
      
      # If no data, return empty result
      if df.empty:
          return jsonify({'dates': [], 'planThickness': [], 'finalThk': [], 'planWidth': [], 'finalWidth': []})
      
      # Convert clearanceDate to datetime and sort
      df['clearanceDate'] = pd.to_datetime(df['clearanceDate'])
      df = df.sort_values('clearanceDate')
      
      # Convert dates to strings for JSON serialization
      dates = [d.strftime('%Y-%m-%d %H:%M:%S') for d in df['clearanceDate']]
      
      return jsonify({
          'dates': dates,
          'planThickness': df['planThickness'].tolist(),
          'finalThk': df['finalThk'].tolist(),
          'planWidth': df['planWidth'].tolist(),
          'finalWidth': df['finalWidth'].tolist(),
          'products': df['product'].tolist() if 'product' in df.columns else []
      })
  except Exception as e:
      return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/visualization_3d')
def get_3d_visualization():
  if db is None:
      return jsonify({'error': 'Database connection error'}), 500
      
  try:
      # Get parameters for 3D visualization
      x_axis = request.args.get('x_axis', 'finalThk')
      y_axis = request.args.get('y_axis', 'finalWidth')
      z_axis = request.args.get('z_axis', 'finalWeight')
      color_by = request.args.get('color_by', 'product')
      visualization_type = request.args.get('visualization_type', 'scatter')
      from_date = request.args.get('from_date')
      to_date = request.args.get('to_date')
      
      # Build query
      query = {}
      if from_date and to_date:
          query['clearanceDate'] = {
              '$gte': from_date,
              '$lte': to_date
          }
      elif from_date:
          query['clearanceDate'] = {'$gte': from_date}
      elif to_date:
          query['clearanceDate'] = {'$lte': to_date}
      
      # Fetch coil data
      projection = {'_id': 0}
      
      # Make sure we include the required fields
      for field in [x_axis, y_axis, z_axis, color_by, 'serialNo', 'product', 'finalStatus', 'clearanceDate']:
          if field:
              projection[field] = 1
      
      # Fetch all data matching the query
      coil_data = list(db.coil_production.find(query, projection))
      
      # Generate 3D visualization
      visualization_3d = generate_3d_visualization(
          coil_data, 
          x_axis=x_axis, 
          y_axis=y_axis, 
          z_axis=z_axis, 
          color_by=color_by,
          visualization_type=visualization_type
      )
      
      return jsonify({'visualization_3d': visualization_3d})
  except Exception as e:
      return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/table_columns')
def get_table_columns():
    try:
        # Define column definitions
        columns = [
            {'field': 'slNo', 'title': 'SL No'},
            {'field': 'product', 'title': 'Product'},
            {'field': 'serialNo', 'title': 'Serial No'},
            {'field': 'motherSerialNo', 'title': 'Mother Serial No'},
            {'field': 'planThickness', 'title': 'Plan Thickness'},
            {'field': 'finalThk', 'title': 'Final Thickness'},
            {'field': 'planWidth', 'title': 'Plan Width'},
            {'field': 'finalWidth', 'title': 'Final Width'},
            {'field': 'planTdc', 'title': 'Plan TDC'},
            {'field': 'finalTdc', 'title': 'Final TDC'},
            {'field': 'ipIdm', 'title': 'IP IDM'},
            {'field': 'orderCustomer', 'title': 'Order Customer'},
            {'field': 'actualCustomer', 'title': 'Actual Customer'},
            {'field': 'orderPath', 'title': 'Order Path'},
            {'field': 'actualPath', 'title': 'Actual Path'},
            {'field': 'orderWeightMinMax', 'title': 'Order Weight Min-Max'},
            {'field': 'finalWeight', 'title': 'Final Weight'},
            {'field': 'orderYieldStrength', 'title': 'Order Yield Strength'},
            {'field': 'yieldStrength', 'title': 'Yield Strength'},
            {'field': 'orderElongation', 'title': 'Order Elongation'},
            {'field': 'elongation', 'title': 'Elongation'},
            {'field': 'orderHardness', 'title': 'Order Hardness'},
            {'field': 'hardness', 'title': 'Hardness'},
            {'field': 'linerMarking', 'title': 'Liner Marking'},
            {'field': 'sleeveType', 'title': 'Sleeve Type'},
            {'field': 'labResult', 'title': 'Lab Result'},
            {'field': 'labTestRemark', 'title': 'Lab Test Remark'},
            {'field': 'surfaceResult', 'title': 'Surface Result'},
            {'field': 'surfaceRemark', 'title': 'Surface Remark'},
            {'field': 'finalStatus', 'title': 'Final Status'},
            {'field': 'majorDefect', 'title': 'Major Defect'},
            {'field': 'defAllocation', 'title': 'Def Allocation'},
            {'field': 'reworkUnit', 'title': 'Rework Unit'},
            {'field': 'suggestions', 'title': 'Suggestions'},
            {'field': 'prevUnit', 'title': 'Prev Unit'},
            {'field': 'ncoFlag', 'title': 'NCO Flag'},
            {'field': 'deversionCat', 'title': 'Deversion Cat'},
            {'field': 'clearanceDate', 'title': 'Clearance Date'},
            {'field': 'age', 'title': 'Age'},
            {'field': 'material', 'title': 'Material'},
            {'field': 'oilType', 'title': 'Oil Type'},
            {'field': 'oilerUsage', 'title': 'Oiler Usage'},
            {'field': 'shift', 'title': 'Shift'}
        ]
        
        # Define default visible columns (a subset for better performance)
        default_visible = [
            'slNo', 'product', 'serialNo', 'finalThk', 'finalWidth', 
            'finalWeight', 'finalStatus', 'clearanceDate', 'age'
        ]
        
        return jsonify({
            'columns': columns,
            'defaultVisible': default_visible
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/table_data')
def get_table_data():
    if db is None:
        return jsonify({'error': 'Database connection error'}), 500
        
    try:
        # Get filter parameters
        serial_no = request.args.get('serial_no')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        product = request.args.get('product')
        status = request.args.get('status')
        
        # Get pagination parameters
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        
        # Get sorting parameters
        sort_by = request.args.get('sort_by', 'slNo')
        sort_direction = request.args.get('sort_direction', 'asc')
        
        # Build query
        query = {}
        if from_date and to_date:
            query['clearanceDate'] = {
                '$gte': from_date,
                '$lte': to_date
            }
        if serial_no:
            query['serialNo'] = {'$regex': serial_no, '$options': 'i'}
        if product:
            query['product'] = product
        if status:
            query['finalStatus'] = status
        
        # Count total documents matching the query
        total = db.coil_production.count_documents(query)
        
        # Calculate skip value for pagination
        skip = (page - 1) * limit
        
        # Determine sort direction for MongoDB
        mongo_sort_direction = 1 if sort_direction == 'asc' else -1
        
        # Fetch paginated and sorted data
        data = list(db.coil_production.find(
            query, 
            {'_id': 0}
        ).sort(
            sort_by, 
            mongo_sort_direction
        ).skip(skip).limit(limit))
        
        return jsonify({
            'data': json.loads(json_util.dumps(data)),
            'total': total,
            'page': page,
            'limit': limit,
            'pages': math.ceil(total / limit)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/filter_options')
def get_filter_options():
    if db is None:
        return jsonify({'error': 'Database connection error'}), 500
        
    try:
        # Get unique products
        products = db.coil_production.distinct('product')
        
        # Get unique statuses
        statuses = db.coil_production.distinct('finalStatus')
        
        return jsonify({
            'products': products,
            'statuses': statuses
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@production_overview.route('/export_data')
def export_data():
    try:
        from_date = request.args.get('from_date', None)
        to_date = request.args.get('to_date', None)
        serial_no = request.args.get('serial_no', None)
        product = request.args.get('product', None)
        status = request.args.get('status', None)
        export_type = request.args.get('type', 'table')
        
        # Get sorting parameters
        sort_by = request.args.get('sort_by', None)
        sort_direction = request.args.get('sort_direction', None)
        
        # Get 3D visualization parameters if applicable
        x_axis = request.args.get('x_axis', 'finalThk')
        y_axis = request.args.get('y_axis', 'finalWidth')
        z_axis = request.args.get('z_axis', 'finalWeight')
        color_by = request.args.get('color_by', 'product')
        visualization_type = request.args.get('visualization_type', 'scatter')

        # Build query
        query = {}
        if from_date and to_date:
            query['clearanceDate'] = {
                '$gte': from_date,
                '$lte': to_date
            }
        if serial_no:
            query['serialNo'] = {'$regex': serial_no, '$options': 'i'}
        if product:
            query['product'] = product
        if status:
            query['finalStatus'] = status
        
        # Only apply sorting if sort parameters are provided
        if sort_by and sort_direction and export_type == 'table':
            # Determine sort direction for MongoDB
            mongo_sort_direction = 1 if sort_direction == 'asc' else -1
            # Use MongoDB's sorting
            data = list(db.coil_production.find(query, {'_id': 0}).sort(sort_by, mongo_sort_direction))
        else:
            # No sorting applied, just fetch the data
            data = list(db.coil_production.find(query, {'_id': 0}))
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # If DataFrame is empty, add a message
        if df.empty:
            df = pd.DataFrame([{'Message': 'No data found for the selected filters'}])
        
        # Create Excel file in memory
        output = BytesIO()
        
        # Use openpyxl engine
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Get the workbook and worksheet
            workbook = writer.book
            
            # Create a new worksheet for the data
            worksheet = workbook.create_sheet("Data")
            
            # Add title row
            worksheet.cell(row=1, column=1, value="Coil Production Data")
            worksheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
            
            # Track the current row for dynamic positioning
            current_row = 2
            
            # Add filter information rows - each filter on a separate row
            if from_date and to_date:
                worksheet.cell(row=current_row, column=1, value=f"Date Range: {from_date} to {to_date}")
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            elif from_date:
                worksheet.cell(row=current_row, column=1, value=f"Date From: {from_date}")
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            elif to_date:
                worksheet.cell(row=current_row, column=1, value=f"Date To: {to_date}")
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            
            if serial_no:
                worksheet.cell(row=current_row, column=1, value=f"Serial No: {serial_no}")
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            
            if product:
                worksheet.cell(row=current_row, column=1, value=f"Product: {product}")
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            
            if status:
                worksheet.cell(row=current_row, column=1, value=f"Status: {status}")
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            
            # Add sorting information row only for table exports with explicit sorting
            if export_type == 'table' and sort_by and sort_direction:
                sort_info = f"Sorted by: {sort_by} ({sort_direction})"
                worksheet.cell(row=current_row, column=1, value=sort_info)
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            
            # Add 3D visualization parameters if applicable
            if export_type == '3d':
                viz_info = f"3D Visualization: X-Axis={x_axis}, Y-Axis={y_axis}, Z-Axis={z_axis}, Color By={color_by}, Type={visualization_type}"
                worksheet.cell(row=current_row, column=1, value=viz_info)
                worksheet.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
                current_row += 1
            
            # Add an empty row between filter/sort info and table data
            current_row += 1
            
            # Write column headers
            header_row = current_row
            for c_idx, column in enumerate(df.columns, 1):
                worksheet.cell(row=header_row, column=c_idx, value=column)
                worksheet.cell(row=header_row, column=c_idx).font = openpyxl.styles.Font(bold=True)
            
            # Write the DataFrame starting from the row after headers
            for r_idx, row in enumerate(df.iterrows(), header_row + 1):
                for c_idx, value in enumerate(row[1], 1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Remove the default sheet
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]
        
        output.seek(0)
        
        # Generate filename with current timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'coil_production_{export_type}_{timestamp}.xlsx'
        )
    except Exception as e:
        print(f"Export error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@production_overview.route('/production_overview/coil_distribution_chart')
def get_coil_distribution_chart():
  if db is None:
      return jsonify({'error': 'Database connection error'}), 500
      
  try:
      # Fetch coil data
      coil_data = list(db.coil_production.find({}, {'_id': 0, 'product': 1}))
      
      # Process data for chart
      df = pd.DataFrame(coil_data)
      
      # If no data, return empty result
      if df.empty:
          return jsonify({'labels': [], 'data': []})
      
      # Count products
      product_counts = df['product'].value_counts().reset_index()
      product_counts.columns = ['product', 'count']
      
      # Calculate percentages
      total_products = product_counts['count'].sum()
      product_counts['percentage'] = (product_counts['count'] / total_products * 100).round(1)
      
      # Sort by count in descending order
      product_counts = product_counts.sort_values('count', ascending=False)
      
      return jsonify({
          'labels': product_counts['product'].tolist(),
          'data': product_counts['count'].tolist(),
          'percentages': product_counts['percentage'].tolist()
      })
  except Exception as e:
      return jsonify({'error': str(e)}), 500

def generate_3d_visualization(data, x_axis='finalThk', y_axis='finalWidth', z_axis='finalWeight', color_by='product', visualization_type='scatter'):
    """
    Generate a 3D visualization of the coil data.
    
    Parameters:
    - data: List of dictionaries containing coil data
    - x_axis: Parameter to use for X-axis
    - y_axis: Parameter to use for Y-axis
    - z_axis: Parameter to use for Z-axis
    - color_by: Parameter to use for coloring points
    - visualization_type: Type of visualization ('scatter', 'surface', 'mesh')
    
    Returns:
    - HTML representation of the 3D visualization
    """
    if not data:
        return "<p class='text-gray-500'>No data available for visualization</p>"
    
    df = pd.DataFrame(data)
    
    # Ensure all selected parameters are numeric
    for param in [x_axis, y_axis, z_axis]:
        if param in df.columns:
            df[param] = pd.to_numeric(df[param], errors='coerce')
    
    # Drop rows with missing values in the selected parameters
    df = df.dropna(subset=[x_axis, y_axis, z_axis])
    
    if df.empty:
        return "<p class='text-gray-500'>No valid data available for the selected parameters</p>"
    
    # Create the appropriate 3D visualization based on the type
    if visualization_type == 'scatter':
        fig = create_3d_scatter(df, x_axis, y_axis, z_axis, color_by)
    elif visualization_type == 'surface':
        fig = create_3d_surface(df, x_axis, y_axis, z_axis, color_by)
    elif visualization_type == 'mesh':
        fig = create_3d_mesh(df, x_axis, y_axis, z_axis, color_by)
    else:
        # Default to scatter if an invalid type is provided
        fig = create_3d_scatter(df, x_axis, y_axis, z_axis, color_by)
    
    # Update layout for better appearance
    fig.update_layout(
        title=f'3D Visualization of Coil Parameters',
        scene=dict(
            xaxis_title=x_axis,
            yaxis_title=y_axis,
            zaxis_title=z_axis,
            aspectmode='cube'
        ),
        margin=dict(l=0, r=0, b=0, t=40),
        height=600
    )
    
    return fig.to_html(full_html=False, include_plotlyjs='cdn')

def create_3d_scatter(df, x_axis, y_axis, z_axis, color_by):
    """Create a 3D scatter plot."""
    if color_by in df.columns:
        # If color_by is categorical
        if df[color_by].dtype == 'object' or len(df[color_by].unique()) < 10:
            fig = px.scatter_3d(
                df, 
                x=x_axis, 
                y=y_axis, 
                z=z_axis,
                color=color_by,
                hover_name='serialNo' if 'serialNo' in df.columns else None,
                hover_data=['product', 'finalStatus'] if all(col in df.columns for col in ['product', 'finalStatus']) else None,
                opacity=0.7,
                size_max=10
            )
        # If color_by is numerical
        else:
            fig = px.scatter_3d(
                df, 
                x=x_axis, 
                y=y_axis, 
                z=z_axis,
                color=color_by,
                color_continuous_scale=px.colors.sequential.Viridis,
                hover_name='serialNo' if 'serialNo' in df.columns else None,
                hover_data=['product', 'finalStatus'] if all(col in df.columns for col in ['product', 'finalStatus']) else None,
                opacity=0.7,
                size_max=10
            )
    else:
        # If color_by column doesn't exist, use a default color
        fig = px.scatter_3d(
            df, 
            x=x_axis, 
            y=y_axis, 
            z=z_axis,
            hover_name='serialNo' if 'serialNo' in df.columns else None,
            hover_data=['product', 'finalStatus'] if all(col in df.columns for col in ['product', 'finalStatus']) else None,
            opacity=0.7,
            size_max=10
        )
    
    return fig

def create_3d_surface(df, x_axis, y_axis, z_axis, color_by):
    """Create a 3D surface plot."""
    try:
        # For surface plots, we need to create a grid of values
        # First, create a 2D grid of x and y values
        x_range = np.linspace(df[x_axis].min(), df[x_axis].max(), 20)
        y_range = np.linspace(df[y_axis].min(), df[y_axis].max(), 20)
        x_grid, y_grid = np.meshgrid(x_range, y_range)
        
        try:
            # Try to import scipy
            from scipy.interpolate import griddata
            
            # Interpolate z values for the grid
            z_grid = griddata(
                (df[x_axis], df[y_axis]), 
                df[z_axis], 
                (x_grid, y_grid), 
                method='cubic', 
                fill_value=df[z_axis].mean()
            )
            
            # Create the surface plot
            fig = go.Figure(data=[go.Surface(
                x=x_grid, 
                y=y_grid, 
                z=z_grid,
                colorscale='Viridis',
                opacity=0.8
            )])
            
            return fig
        except ImportError:
            # If scipy is not available, fall back to a simpler 3D scatter plot
            print("scipy.interpolate not available, falling back to scatter plot")
            return create_3d_scatter(df, x_axis, y_axis, z_axis, color_by)
    except Exception as e:
        print(f"Error creating 3D surface: {str(e)}")
        return create_3d_scatter(df, x_axis, y_axis, z_axis, color_by)

def create_3d_mesh(df, x_axis, y_axis, z_axis, color_by):
    """Create a 3D mesh plot."""
    try:
        # For mesh plots, we'll use a similar approach to surface plots
        # but with a different visualization
        x_range = np.linspace(df[x_axis].min(), df[x_axis].max(), 15)
        y_range = np.linspace(df[y_axis].min(), df[y_axis].max(), 15)
        x_grid, y_grid = np.meshgrid(x_range, y_range)
        
        try:
            # Try to import scipy
            from scipy.interpolate import griddata
            
            # Interpolate z values for the grid
            z_grid = griddata(
                (df[x_axis], df[y_axis]), 
                df[z_axis], 
                (x_grid, y_grid), 
                method='cubic', 
                fill_value=df[z_axis].mean()
            )
            
            # Create the mesh plot
            fig = go.Figure(data=[go.Mesh3d(
                x=x_grid.flatten(),
                y=y_grid.flatten(),
                z=z_grid.flatten(),
                intensity=z_grid.flatten(),
                colorscale='Viridis',
                opacity=0.8
            )])
            
            return fig
        except ImportError:
            # If scipy is not available, fall back to a simpler 3D scatter plot
            print("scipy.interpolate not available, falling back to scatter plot")
            return create_3d_scatter(df, x_axis, y_axis, z_axis, color_by)
    except Exception as e:
        print(f"Error creating 3D mesh: {str(e)}")
        return create_3d_scatter(df, x_axis, y_axis, z_axis, color_by)

def generate_coil_distribution_chart(coil_data):
  """
  Generates a pie chart showing the distribution of coil products (GI, GL, etc.).
  """
  if not coil_data:
      return "<p class='text-gray-500'>No coil product data available</p>"

  df = pd.DataFrame(coil_data)
  
  # Count products
  product_counts = df['product'].value_counts().reset_index()
  product_counts.columns = ['product', 'count']
  
  # Calculate percentages
  total_products = product_counts['count'].sum()
  product_counts['percentage'] = (product_counts['count'] / total_products * 100).round(1)
  
  # Sort by count in descending order
  product_counts = product_counts.sort_values('count', ascending=False)
  
  # Use a colorful palette
  colors = [
      'rgb(82, 113, 255)',   # Blue
      'rgb(255, 99, 71)',    # Coral
      'rgb(34, 197, 94)',    # Green
      'rgb(168, 85, 247)',   # Purple
      'rgb(251, 146, 60)',   # Orange
      'rgb(236, 72, 153)',   # Pink
      'rgb(234, 179, 8)'     # Yellow
  ]

  fig = go.Figure(data=[go.Pie(
      labels=product_counts['product'],
      values=product_counts['count'],
      hole=0.4,
      marker=dict(colors=colors),
      textinfo='label+percent',
      hovertemplate="<b>%{label}</b><br>" +
                   "Count: %{value}<br>" +
                   "Percentage: %{percent}<br>" +
                   "<extra></extra>",
      texttemplate="%{percent}",
      textposition='outside',
      textfont=dict(size=10)
  )])

  fig.update_layout(
      title=dict(
          text='Coil Product Distribution',
          x=0.5,
          y=0.95,
          xanchor='center',
          yanchor='top',
          font=dict(size=16)
      ),
      showlegend=False,
      height=300,
      margin=dict(l=20, r=20, t=40, b=20),
      plot_bgcolor='white'
  )

  return fig.to_html(
      full_html=False,
      include_plotlyjs='cdn',
      config={
          'displayModeBar': False,
          'responsive': True
      }
  )

def generate_daily_production_chart(coil_data):
  """
  Generates a daily production chart showing the number of coils produced each day.
  """
  if not coil_data:
      return "<p class='text-gray-500'>No daily production data available</p>"

  df = pd.DataFrame(coil_data)
  
  # Convert clearanceDate to datetime
  df['clearanceDate'] = pd.to_datetime(df['clearanceDate'])
  
  # Extract date from clearanceDate (without time)
  df['production_date'] = df['clearanceDate'].dt.date
  
  # Group by date and count coils
  daily_counts = df.groupby('production_date').size().reset_index(name='count')
  daily_counts = daily_counts.sort_values('production_date')
  
  # Create the bar chart
  fig = go.Figure()
  fig.add_trace(go.Bar(
      x=daily_counts['production_date'],
      y=daily_counts['count'],
      marker_color='rgb(55, 83, 109)',
      text=daily_counts['count'],
      textposition='auto',
      hovertemplate='Date: %{x}<br>Number of Coils: %{y}<extra></extra>'
  ))
  
  fig.update_layout(
      title='Daily Coil Production',
      xaxis_title='Date',
      yaxis_title='Number of Coils',
      plot_bgcolor='rgba(242, 244, 247, 1)',
      paper_bgcolor='white',
      height=300,
      margin=dict(l=40, r=40, t=40, b=40)
  )
  
  # If there are too many dates, limit the number of ticks shown
  if len(daily_counts) > 15:
      # Show every nth date to avoid overcrowding
      n = max(1, len(daily_counts) // 15)
      tickvals = daily_counts['production_date'].iloc[::n]
      fig.update_xaxes(
          tickangle=45,
          tickmode='array',
          tickvals=tickvals,
          showgrid=True, 
          gridwidth=1, 
          gridcolor='white'
      )
  else:
      fig.update_xaxes(
          tickangle=45,
          tickmode='array',
          tickvals=daily_counts['production_date'],
          showgrid=True, 
          gridwidth=1, 
          gridcolor='white'
      )
  
  fig.update_yaxes(
      showgrid=True, 
      gridwidth=1, 
      gridcolor='white'
  )
  
  return fig.to_html(full_html=False)

def generate_coil_analysis_chart(coil_data):
    """
    Generates a detailed analysis chart for coil production data.
    """
    if not coil_data:
        return "<p class='text-gray-500'>No coil analysis data available</p>"

    df = pd.DataFrame(coil_data)
    
    # Convert clearanceDate to datetime and sort
    df['clearanceDate'] = pd.to_datetime(df['clearanceDate'])
    df = df.sort_values('clearanceDate')
    
    fig = go.Figure()

    # Thickness comparison
    fig.add_trace(go.Scatter(
        x=df['clearanceDate'],
        y=df['planThickness'],
        name='Planned Thickness',
        line=dict(color='#1f77b4', width=2),
        mode='lines+markers',
        hovertemplate="<b>Serial No:</b> %{customdata}<br>" +
                     "<b>Date:</b> %{x|%Y-%m-%d %H:%M:%S}<br>" +
                     "<b>Planned Thickness:</b> %{y:.2f} mm<extra></extra>",
        customdata=df['product'] if 'product' in df.columns else None
    ))
    fig.add_trace(go.Scatter(
        x=df['clearanceDate'],
        y=df['finalThk'],
        name='Final Thickness',
        line=dict(color='#ff7f0e', width=2),
        mode='lines+markers',
        hovertemplate="<b>Serial No:</b> %{customdata}<br>" +
                     "<b>Date:</b> %{x|%Y-%m-%d %H:%M:%S}<br>" +
                     "<b>Final Thickness:</b> %{y:.2f} mm<extra></extra>",
        customdata=df['product'] if 'product' in df.columns else None
    ))

    # Width comparison
    fig.add_trace(go.Scatter(
        x=df['clearanceDate'],
        y=df['planWidth'],
        name='Planned Width',
        line=dict(color='#2ca02c', width=2),
        mode='lines+markers',
        yaxis='y2',
        hovertemplate="<b>Serial No:</b> %{customdata}<br>" +
                     "<b>Date:</b> %{x|%Y-%m-%d %H:%M:%S}<br>" +
                     "<b>Planned Width:</b> %{y:.0f} mm<extra></extra>",
        customdata=df['product'] if 'product' in df.columns else None
    ))
    fig.add_trace(go.Scatter(
        x=df['clearanceDate'],
        y=df['finalWidth'],
        name='Final Width',
        line=dict(color='#d62728', width=2),
        mode='lines+markers',
        yaxis='y2',
        hovertemplate="<b>Serial No:</b> %{customdata}<br>" +
                     "<b>Date:</b> %{x|%Y-%m-%d %H:%M:%S}<br>" +
                     "<b>Final Width:</b> %{y:.0f} mm<extra></extra>",
        customdata=df['product'] if 'product' in df.columns else None
    ))

    # Customize the layout
    fig.update_layout(
        title='Coil Production Analysis',
        xaxis=dict(
            title='Date and Time',
            tickangle=45,
            tickformat='%Y-%m-%d',
            tickfont=dict(size=10),
            type='date'
        ),
        yaxis=dict(
            title='Thickness (mm)',
            side='left',
            showgrid=True,
            gridcolor='rgba(0,0,0,0.1)'
        ),
        yaxis2=dict(
            title='Width (mm)',
            side='right',
            overlaying='y',
            showgrid=False,
            gridcolor='rgba(0,0,0,0.1)'
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        ),
        hovermode='x unified',
        plot_bgcolor='rgba(240, 240, 240, 0.8)',
        height=500,
        margin=dict(l=50, r=50, t=80, b=50)
    )

    return fig.to_html(full_html=False, include_plotlyjs='cdn')






